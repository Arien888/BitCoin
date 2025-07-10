import yaml
from pathlib import Path
from openpyxl import load_workbook, Workbook
import win32com.client  # ← Excel制御用（Windows専用）

# ✅ Excelクエリ更新＆再計算関数
def update_excel_queries_and_recalculate(file_path):
    excel = win32com.client.Dispatch("Excel.Application")
    excel.Visible = False
    excel.DisplayAlerts = False

    wb = excel.Workbooks.Open(str(file_path))

    wb.RefreshAll()
    excel.CalculateUntilAsyncQueriesDone()
    excel.CalculateFullRebuild()

    wb.Save()
    wb.Close(False)
    excel.Quit()

    print(f"✅ Excelクエリ更新・再計算・保存 完了: {file_path}")

# ✅ 値のみコピーする関数（複数シート）
def export_values_only(src_file, sheet_names, dst_file):
    src_wb = load_workbook(src_file, data_only=True)
    new_wb = Workbook()
    first = True

    for name in sheet_names:
        if name not in src_wb.sheetnames:
            print(f"⚠ シート '{name}' は存在しません。スキップします。")
            continue

        src_ws = src_wb[name]
        if first:
            new_ws = new_wb.active
            new_ws.title = name
            first = False
        else:
            new_ws = new_wb.create_sheet(title=name)

        for row in src_ws.iter_rows():
            for cell in row:
                new_ws[cell.coordinate].value = cell.value

    new_wb.save(dst_file)
    print(f"✅ エクスポート完了: {dst_file}")

# ✅ メイン処理
def main():
    config_path = Path(__file__).resolve().parent.parent / "config.yaml"
    with open(config_path, "r", encoding="utf-8") as f:
        config = yaml.safe_load(f)

    order_cfg = config["order_export"]
    src_file = (config_path.parent / order_cfg["source_file"]).resolve()
    dst_file = (config_path.parent / order_cfg["output_file"]).resolve()

    # sheet_names取得
    if "sheet_names" in order_cfg:
        sheet_names = order_cfg["sheet_names"]
    elif "sheet_name" in order_cfg:
        sheet_names = [order_cfg["sheet_name"]]
    else:
        raise KeyError("YAMLに 'sheet_name' または 'sheet_names' が見つかりません。")

    # エクスポートと更新処理
    export_values_only(src_file, sheet_names, dst_file)
    update_excel_queries_and_recalculate(dst_file)

# ✅ 実行
if __name__ == "__main__":
    main()
