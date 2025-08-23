import time
import hmac
import hashlib
import base64
import requests
import json
import yaml
import os
import openpyxl

# =========================
# 設定
# =========================
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
with open(os.path.join(BASE_DIR, "..", "config.yaml"), "r", encoding="utf-8") as f:
    config = yaml.safe_load(f)

API_KEY = config["bitget"]["api_key"]
API_SECRET = config["bitget"]["api_secret"]
PASSPHRASE = config["bitget"]["passphrase"]
BASE_URL = "https://api.bitget.com"

# Excel関連
excel_rel_path = config["order_export"]["source_file"]
excel_path = os.path.join(BASE_DIR, "..", excel_rel_path)
sheet_names = [
    config["excel"]["sheets"]["bitget_earn_to_spot"],
]

# =========================
# 共通関数
# =========================
def get_timestamp():
    return str(int(time.time() * 1000))

def sign_bitget(api_secret, timestamp, method, request_path, body=""):
    message = f"{timestamp}{method.upper()}{request_path}{body}"
    signature = hmac.new(api_secret.encode(), message.encode(), hashlib.sha256).digest()
    return base64.b64encode(signature).decode()

def safe_json(response):
    try:
        return response.json()
    except Exception:
        print("JSON変換失敗:", response.text)
        return None

def get_earning_products(coin):
    """特定コインのEarning商品(productId)を取得"""
    url = f"{BASE_URL}/api/v2/earn/savings/list?coin={coin}"
    request_path = f"/api/v2/earn/savings/list?coin={coin}"
    method = "GET"
    timestamp = get_timestamp()
    sign = sign_bitget(API_SECRET, timestamp, method, request_path)
    headers = {
        "ACCESS-KEY": API_KEY,
        "ACCESS-SIGN": sign,
        "ACCESS-TIMESTAMP": timestamp,
        "ACCESS-PASSPHRASE": PASSPHRASE,
    }
    response = requests.get(url, headers=headers)
    print("Earning商品取得ステータス:", response.status_code)
    data = safe_json(response)
    if data and "data" in data:
        for item in data["data"]:
            if item["coin"] == coin:
                return item["productId"], item.get("periodType", "FLEXIBLE")
    return None, None

def redeem_earning(product_id, amount, period_type="FLEXIBLE"):
    """Earning資産を償還"""
    url = f"{BASE_URL}/api/v2/earn/savings/redeem"
    request_path = "/api/v2/earn/savings/redeem"
    method = "POST"
    body_dict = {
        "productId": product_id,
        "amount": str(amount),
        "periodType": period_type.lower(),  # "flexible" or "fixed"
    }
    body = json.dumps(body_dict, separators=(",", ":"))
    timestamp = get_timestamp()
    sign = sign_bitget(API_SECRET, timestamp, method, request_path, body)
    headers = {
        "ACCESS-KEY": API_KEY,
        "ACCESS-SIGN": sign,
        "ACCESS-TIMESTAMP": timestamp,
        "ACCESS-PASSPHRASE": PASSPHRASE,
        "Content-Type": "application/json",
    }
    response = requests.post(url, headers=headers, data=body)
    print(f"償還リクエスト: {product_id} {amount} {period_type} → ステータス {response.status_code}")
    print(response.text)
    return safe_json(response)

# =========================
# 実行処理
# =========================
if __name__ == "__main__":
    # Excel読込
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    for sheet_name in sheet_names:
        if sheet_name not in wb.sheetnames:
            print(f"シート {sheet_name} が存在しません")
            continue

        sheet = wb[sheet_name]
        print(f"シート {sheet_name} 読込中...")

        # 1行目はヘッダ想定
        for row in sheet.iter_rows(min_row=2, values_only=True):
            coin, amount, period_type = row[:3]
            if not coin or not amount:
                continue

            print(f"注文読込: coin={coin}, amount={amount}, periodType={period_type}")

            # productIdをAPIから取得
            product_id, prod_period = get_earning_products(coin)
            if not product_id:
                print(f"❌ {coin} の productId 取得失敗 → スキップ")
                continue

            # 償還実行
            redeem_earning(product_id, amount, period_type or prod_period)
            time.sleep(1)  # API制限対策
