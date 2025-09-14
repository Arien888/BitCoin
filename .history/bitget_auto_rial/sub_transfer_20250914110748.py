import sys
import io
import os
import yaml
import time
import requests
import hmac
import hashlib
import base64
import json
import openpyxl  # Excel読み込み用

# --- 出力をUTF-8に設定 ---
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")

# --- スクリプトのあるディレクトリ ---
BASE_DIR = os.path.dirname(os.path.abspath(__file__))

# --- config.yaml 読み込み ---
with open(os.path.join(BASE_DIR, "..", "config.yaml"), "r", encoding="utf-8") as f:
    config = yaml.safe_load(f)

API_KEY = config["bitget"]["api_key"]
API_SECRET = config["bitget"]["api_secret"]
PASSPHRASE = config["bitget"]["passphrase"]
IS_TESTNET = config["bitget"].get("is_testnet", False)

BASE_URL = "https://api.bitget.com" if not IS_TESTNET else "https://api-testnet.bitget.com"

# --- UIDの設定（config.yamlに記載しておく） ---
MAIN_UID = str(config["bitget"]["main_uid"])
SUB_UID = str(config["bitget"]["subaccount"]["sub_uid"])

# --- Excel 設定 ---
excel_rel_path = config["order_export"]["source_file"]
excel_path = os.path.join(BASE_DIR, "..", excel_rel_path)
sheet_name = config["excel"]["sheets"]["bitget_sub_transfer"]

# --- Bitget API署名生成 ---
def sign_request(method, request_path, body=""):
    timestamp = str(time.time())
    pre_sign = f"{timestamp}{method}{request_path}{body}"
    sign = hmac.new(API_SECRET.encode("utf-8"), pre_sign.encode("utf-8"), hashlib.sha256).digest()
    sign_b64 = base64.b64encode(sign).decode()
    return timestamp, sign_b64

# --- サブアカウント振替 ---
def transfer_between_main_sub(from_uid, to_uid, from_type, to_type, coin, amount, client_oid=None):
    path = "/api/v2/spot/wallet/subaccount-transfer"
    url = BASE_URL + path

    body = {
        "fromType": from_type,
        "toType": to_type,
        "coin": coin,
        "amount": str(amount),
        "fromUserId": str(from_uid),
        "toUserId": str(to_uid),
        "clientOid": client_oid or str(int(time.time() * 1000))
    }

    body_str = json.dumps(body, separators=(",", ":"))
    timestamp, sign = sign_request("POST", path, body_str)

    headers = {
        "ACCESS-KEY": API_KEY,
        "ACCESS-SIGN": sign,
        "ACCESS-TIMESTAMP": timestamp,
        "ACCESS-PASSPHRASE": PASSPHRASE,
        "Content-Type": "application/json"
    }

    r = requests.post(url, headers=headers, data=body_str)
    return r.json()

# --- Excel から金額取得 ---
def read_transfer_amounts():
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Excelにシート {sheet_name} が存在しません")
    ws = wb[sheet_name]

    main_to_sub = ws["A2"].value or 0
    sub_to_main = ws["B2"].value or 0

    return float(main_to_sub), float(sub_to_main)

# --- メイン処理 ---
def main():
    try:
        print("[INFO] Excelから振替金額を読み込みます...")
        main_to_sub, sub_to_main = read_transfer_amounts()
        print(f"[INFO] メイン→サブ先物: {main_to_sub}, サブ先物→メイン: {sub_to_main}")

        if main_to_sub > 0:
            print(f"[INFO] メイン Spot → サブUSDT先物 に {main_to_sub} USDT を送金します...")
            result = transfer_between_main_sub(
                MAIN_UID, SUB_UID,
                "spot", "usdt_futures",
                "USDT", main_to_sub
            )
            print("[INFO] 振替結果(メイン→サブ):")
            print(result)

        if sub_to_main > 0:
            print(f"[INFO] サブUSDT先物 → メイン Spot に {sub_to_main} USDT を送金します...")
            result = transfer_between_main_sub(
                SUB_UID, MAIN_UID,
                "usdt_futures", "spot",
                "USDT", sub_to_main
            )
            print("[INFO] 振替結果(サブ→メイン):")
            print(result)

    except Exception as e:
        import traceback
        print(f"[ERROR] 処理中に例外が発生しました: {e}")
        print(traceback.format_exc())

if __name__ == "__main__":
    main()
