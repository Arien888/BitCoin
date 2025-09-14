import sys
import io
import os
import yaml
import time
import requests
import hmac
import hashlib
import base64
import json
import openpyxl  # Excel読み込み用

# --- 出力をUTF-8に設定 ---
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")

# --- スクリプトのあるディレクトリ ---
BASE_DIR = os.path.dirname(os.path.abspath(__file__))

# --- config.yaml 読み込み ---
with open(os.path.join(BASE_DIR, "..", "config.yaml"), "r", encoding="utf-8") as f:
    config = yaml.safe_load(f)

API_KEY = config["bitget"]["api_key"]
API_SECRET = config["bitget"]["api_secret"]
PASSPHRASE = config["bitget"]["passphrase"]
IS_TESTNET = config["bitget"].get("is_testnet", False)

BASE_URL = "https://api.bitget.com" if not IS_TESTNET else "https://api-testnet.bitget.com"

# --- UIDの設定（config.yamlに記載しておく） ---
MAIN_UID = str(config["bitget"]["main_uid"])
SUB_UID = str(config["bitget"]["subaccount"]["sub_uid"])

# --- Excel 設定 ---
excel_rel_path = config["order_export"]["source_file"]
excel_path = os.path.join(BASE_DIR, "..", excel_rel_path)
sheet_name = config["excel"]["sheets"]["bitget_sub_transfer"]

def sign(method, path, body=""):
    """署名を生成"""
    timestamp = str(int(time.time() * 1000))
    message = timestamp + method + path + body
    mac = hmac.new(API_SECRET.encode("utf-8"), message.encode("utf-8"), hashlib.sha256)
    d = mac.digest()
    return timestamp, base64.b64encode(d).decode("utf-8")

def transfer(from_uid, to_uid, from_type, to_type, coin, amount):
    """サブアカウント振替を実行"""
    path = "/api/v2/spot/wallet/subaccount-transfer"
    url = BASE_URL + path

    body_dict = {
        "fromUserId": str(from_uid),
        "toUserId": str(to_uid),
        "fromType": from_type,
        "toType": to_type,
        "coin": coin,
        "amount": str(amount),
        "clientOid": str(uuid.uuid4())  # 一意なID
    }
    body = json.dumps(body_dict)

    timestamp, signature = sign("POST", path, body)

    headers = {
        "ACCESS-KEY": API_KEY,
        "ACCESS-SIGN": signature,
        "ACCESS-TIMESTAMP": timestamp,
        "ACCESS-PASSPHRASE": API_PASSPHRASE,
        "Content-Type": "application/json"
    }

    response = requests.post(url, headers=headers, data=body)
    return response.json()

def main():
    print("[INFO] Excelから振替金額を読み込みます...")
    wb = load_workbook(EXCEL_FILE)
    ws = wb[SHEET_NAME]

    # Excel の 2行目から金額を読み込み
    main_to_sub = ws["A2"].value or 0
    sub_to_main = ws["B2"].value or 0

    print(f"[INFO] メイン→サブ先物: {main_to_sub}, サブ先物→メイン: {sub_to_main}")

    # === メイン → サブ (Spot → USDT先物) ===
    if main_to_sub > 0:
        print(f"[INFO] メイン Spot → サブ USDT先物 に {main_to_sub} USDT を送金します...")
        result = transfer(
            from_uid=MAIN_UID, to_uid=SUB_UID,
            from_type="spot", to_type="usdt_futures",
            coin="USDT", amount=main_to_sub
        )
        print("[INFO] 振替結果(メイン→サブ):")
        print(result)

    # === サブ → メイン (USDT先物 → Spot) ===
    if sub_to_main > 0:
        print(f"[INFO] サブ USDT先物 → メイン Spot に {sub_to_main} USDT を送金します...")
        result = transfer(
            from_uid=SUB_UID, to_uid=MAIN_UID,
            from_type="usdt_futures", to_type="spot",
            coin="USDT", amount=sub_to_main
        )
        print("[INFO] 振替結果(サブ→メイン):")
        print(result)

if __name__ == "__main__":
    import json
    main()