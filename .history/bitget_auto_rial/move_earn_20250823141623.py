import time
import hmac
import hashlib
import base64
import requests
import json
import yaml
import os
import openpyxl  # Excel操作用

# =========================
# 設定
# =========================
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
with open(os.path.join(BASE_DIR, "..", "config.yaml"), "r", encoding="utf-8") as f:
    config = yaml.safe_load(f)

API_KEY = config["bitget"]["api_key"]
API_SECRET = config["bitget"]["api_secret"]
PASSPHRASE = config["bitget"]["passphrase"]
BASE_URL = "https://api.bitget.com"

# Excel設定
sheet_names = [
    config["excel"]["sheets"]["bitget_earn_to_spot"],
]
excel_rel_path = config["order_export"]["source_file"]
excel_path = os.path.join(BASE_DIR, "..", excel_rel_path)


# =========================
# 関数
# =========================
def get_timestamp():
    return str(int(time.time() * 1000))


def sign_bitget(api_secret, timestamp, method, request_path, body=""):
    message = f"{timestamp}{method.upper()}{request_path}{body}"
    signature = hmac.new(api_secret.encode(), message.encode(), hashlib.sha256).digest()
    return base64.b64encode(signature).decode()


def safe_json(response):
    try:
        return response.json()
    except Exception:
        print("JSON変換失敗:", response.text)
        return None


def get_earning_assets():
    url = f"{BASE_URL}/api/v2/earn/account/assets"
    request_path = "/api/v2/earn/account/assets"
    method = "GET"
    timestamp = get_timestamp()
    sign = sign_bitget(API_SECRET, timestamp, method, request_path)
    headers = {
        "ACCESS-KEY": API_KEY,
        "ACCESS-SIGN": sign,
        "ACCESS-TIMESTAMP": timestamp,
        "ACCESS-PASSPHRASE": PASSPHRASE,
        "Content-Type": "application/json",
    }
    response = requests.get(url, headers=headers)
    print("Earning資産取得ステータス:", response.status_code)
    data = safe_json(response)
    if data and "data" in data:
        for asset in data["data"]:
            print("coin:", asset.get("coin"), "amount:", asset.get("amount"))
    return data


def get_earning_product_id(coin, period_type="flexible"):
    url = f"{BASE_URL}/api/v2/earn/savings/product?coin={coin}&filter=available_and_held"
    request_path = f"/api/v2/earn/savings/product?coin={coin}&filter=available_and_held"
    method = "GET"
    timestamp = get_timestamp()
    sign = sign_bitget(API_SECRET, timestamp, method, request_path)
    headers = {
        "ACCESS-KEY": API_KEY,
        "ACCESS-SIGN": sign,
        "ACCESS-TIMESTAMP": timestamp,
        "ACCESS-PASSPHRASE": PASSPHRASE,
        "Content-Type": "application/json",
    }
    response = requests.get(url, headers=headers)
    print("Earning商品取得ステータス:", response.status_code)
    print("レスポンス:", response.text)
    data = safe_json(response)
    if data and "data" in data:
        for item in data["data"]:
            if item.get("coin") == coin and item.get("periodType").lower() == period_type:
                return item.get("productId")
    return None


def redeem_earning(product_id, amount, period_type="flexible"):
    url = f"{BASE_URL}/api/v2/earn/savings/redeem"
    request_path = "/api/v2/earn/savings/redeem"
    method = "POST"
    body_dict = {"productId": product_id, "amount": str(amount), "periodType": period_type}
    body = json.dumps(body_dict, separators=(",", ":"))
    timestamp = get_timestamp()
    sign = sign_bitget(API_SECRET, timestamp, method, request_path, body)
    headers = {
        "ACCESS-KEY": API_KEY,
        "ACCESS-SIGN": sign,
        "ACCESS-TIMESTAMP": timestamp,
        "ACCESS-PASSPHRASE": PASSPHRASE,
        "Content-Type": "application/json",
    }
    response = requests.post(url, headers=headers, data=body)
    print(f"償還リクエスト: {product_id} {amount} → ステータス {response.status_code}")
    print(response.text)
    return safe_json(response)


def read_orders_from_excel(path, sheet_names):
    """Excelから注文データを取得"""
    wb = openpyxl.load_workbook(path, data_only=True)
    orders = []
    for sheet_name in sheet_names:
        if sheet_name not in wb.sheetnames:
            print(f"シート {sheet_name} が存在しません")
            continue
        sheet = wb[sheet_name]
        # ヘッダー想定: Coin | Amount | PeriodType
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if not row[0]:
                continue
            coin = str(row[0]).strip()
            amount = float(row[1]) if row[1] else 0
            period_type = str(row[2]).strip().lower() if row[2] else "flexible"
            orders.append({"coin": coin, "amount": amount, "period_type": period_type})
    return orders


# =========================
# 実行処理
# =========================
if __name__ == "__main__":
    orders = read_orders_from_excel(excel_path, sheet_names)
    print("Excel注文リスト:", orders)

    if not orders:
        print("Excelから注文が取得できませんでした")
        exit()

    assets_data = get_earning_assets()
    if not assets_data or "data" not in assets_data:
        print("Earning資産取得失敗")
        exit()

    # Excelの注文ごとに処理
    for order in orders:
        coin = order["coin"]
        amount = order["amount"]
        period_type = order["period_type"]

        available = 0
        for asset in assets_data["data"]:
            if asset.get("coin") == coin:
                available = float(asset.get("amount", 0))
                break

        if available >= amount:
            product_id = get_earning_product_id(coin, period_type=period_type)
            if product_id:
                print(f"{coin} 利用可能: {available} {coin}, productId: {product_id}")
                redeem_earning(product_id, amount, period_type=period_type)
            else:
                print(f"{coin} の productId が取得できませんでした")
        else:
            print(f"{coin} 残高不足: 要 {amount}, 保有 {available}")
