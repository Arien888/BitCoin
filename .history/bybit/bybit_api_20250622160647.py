import openpyxl

def write_to_existing_excel(excel_path, data_list, keys, sheet_name):
    """
    data_list: list of dict
    keys: list of keys to output as columns
    """
    try:
        wb = openpyxl.load_workbook(excel_path)
    except FileNotFoundError:
        wb = openpyxl.Workbook()

    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        wb.remove(ws)
    ws = wb.create_sheet(title=sheet_name)

    # ヘッダー書き込み
    for col_idx, key in enumerate(keys, start=1):
        ws.cell(row=1, column=col_idx, value=key)

    # データ書き込み
    for row_idx, item in enumerate(data_list, start=2):
        for col_idx, key in enumerate(keys, start=1):
            ws.cell(row=row_idx, column=col_idx, value=item.get(key, ""))

    wb.save(excel_path)
