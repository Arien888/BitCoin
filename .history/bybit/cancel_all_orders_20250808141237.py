import json
import time
import hmac
import hashlib
import requests
from openpyxl import load_workbook
from load_config import load_config  # yamlから読み込む想定

import os

script_dir = os.path.dirname(os.path.abspath(__file__))
cancel_list_excel_path = os.path.join(script_dir, "cancel_list.xlsx")


def generate_signature(api_secret, timestamp, api_key, recv_window, body):
    sign_str = f"{timestamp}{api_key}{recv_window}{body}"
    print(f"[DEBUG] sign_str: {sign_str}")
    signature = hmac.new(
        api_secret.encode(), sign_str.encode(), hashlib.sha256
    ).hexdigest()
    print(f"[DEBUG] signature: {signature}")
    return signature


def cancel_all_orders(api_key, api_secret, symbol, category):
    base_url = "https://api.bybit.com"  # 本番用URL
    path = "/v5/order/cancel-all"
    url = base_url + path

    timestamp = str(int(time.time() * 1000))
    method = "POST"

    body_dict = {"symbol": symbol, "category": category}
    body_json = json.dumps(body_dict, separators=(",", ":"))

    recv_window = "5000"
    timestamp = str(int(time.time() * 1000))
    body_dict = {"symbol": symbol, "category": category}
    body_json = json.dumps(body_dict, separators=(",", ":"))

    signature = generate_signature(
        api_secret, timestamp, api_key, recv_window, body_json
    )

    headers = {
        "Content-Type": "application/json",
        "X-BAPI-API-KEY": api_key,
        "X-BAPI-TIMESTAMP": timestamp,
        "X-BAPI-RECV-WINDOW": "5000",  # 注意：これは署名に含めない
        "X-BAPI-SIGN": signature,
    }

    print(f"[DEBUG] Request URL: {url}")
    print(f"[DEBUG] Request Headers: {headers}")
    print(f"[DEBUG] Request Body: {body_json}")

    response = requests.post(url, headers=headers, data=body_json)
    print(f"[DEBUG] Response Status Code: {response.status_code}")

    print(f"\n=== {symbol}（{category}） キャンセルレスポンス ===")
    print(response.text)

    try:
        return response.json()
    except requests.exceptions.JSONDecodeError:
        print("[ERROR] JSONデコードに失敗しました。レスポンス内容:")
        print(response.text)
        return None


def read_cancel_list_from_excel(
    file_path, sheet_name, symbol_col="A", category_col="B"
):
    wb = load_workbook(file_path, data_only=True)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"シート名 '{sheet_name}' がExcelファイルに存在しません。")
    ws = wb[sheet_name]
    cancel_list = []

    for row in ws.iter_rows(min_row=2):  # 1行目はヘッダー想定
        symbol = row[ord(symbol_col.upper()) - ord("A")].value
        category = row[ord(category_col.upper()) - ord("A")].value
        if symbol and category:
            cancel_list.append((symbol.strip(), category.strip()))
    return cancel_list


def main():
    config = load_config()  # 一つ上の階層のyaml読み込み想定
    api_key = config["bybit"]["key"]
    api_secret = config["bybit"]["secret"]

    sheet_name = "Orders"
    symbol_col = "A"
    category_col = "B"

    cancel_entries = read_cancel_list_from_excel(
        cancel_list_excel_path, sheet_name, symbol_col, category_col
    )

    # for symbol, category in cancel_entries:
    #     cancel_all_orders(api_key, api_secret, symbol, category)
    for symbol, category in cancel_entries:
        cancel_all_orders(api_key, api_secret, symbol, "category")


if __name__ == "__main__":
    main()
