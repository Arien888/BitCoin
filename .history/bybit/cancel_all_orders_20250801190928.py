import json
import time
import hmac
import hashlib
import requests
from load_config import load_config
from openpyxl import load_workbook

def cancel_all_orders(api_key, api_secret, symbol):
    url = "https://api.bybit.com/v5/order/cancel-all"
    timestamp = str(int(time.time() * 1000))
    recv_window = "5000"

    body = {"category": "linear", "symbol": symbol}
    body_json = json.dumps(body, separators=(",", ":"))
    sign_payload = f"{timestamp}{api_key}{recv_window}{body_json}"
    signature = hmac.new(api_secret.encode(), sign_payload.encode(), hashlib.sha256).hexdigest()

    headers = {
        "Content-Type": "application/json",
        "X-BAPI-API-KEY": api_key,
        "X-BAPI-TIMESTAMP": timestamp,
        "X-BAPI-RECV-WINDOW": recv_window,
        "X-BAPI-SIGN": signature,
    }

    response = requests.post(url, headers=headers, data=body_json)
    print(f"=== {symbol} キャンセルレスポンス ===")
    print(response.text)
    return response.json()

def read_symbols_from_excel(file_path, sheet_name, column_letter='A'):
    wb = load_workbook(file_path, data_only=True)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"シート名 '{sheet_name}' がExcelファイルに存在しません。")
    ws = wb[sheet_name]
    symbols = []
    for row in ws[column_letter]:
        val = row.value
        if val and isinstance(val, str):
            symbols.append(val.strip())
    return symbols

def main():
    config = load_config()
    api_key = config["bybit"]["key"]
    api_secret = config["bybit"]["secret"]

    # Excelファイルのパス、シート名を設定
    excel_path = config["excel"]["order_excel_path"]  # 例: "orders.xlsx"
    sheet_name = config["excel"]["order_sheet_name"]  # 例: "Orders"

    symbols = read_symbols_from_excel(excel_path, sheet_name)

    for symbol in symbols:
        cancel_all_orders(api_key, api_secret, symbol)

if __name__ == "__main__":
    main()
