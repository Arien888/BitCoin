import json
import time
import hmac
import hashlib
import requests
from openpyxl import load_workbook
from load_config import load_config  # yamlから読み込む想定


def generate_signature(api_secret, timestamp, method, path, body):
    sign_str = f"{timestamp}{method.upper()}{path}{body}"
    print(f"[DEBUG] sign_str: {sign_str}")  # 署名対象の文字列を出力
    signature = hmac.new(
        api_secret.encode(), sign_str.encode(), hashlib.sha256
    ).hexdigest()
    print(f"[DEBUG] signature: {signature}")  # 生成した署名を出力
    return signature


def cancel_all_orders(api_key, api_secret, symbol, category):
    base_url = "https://api-testnet.bybit.com"  # 本番は https://api.bybit.com
    path = "/v5/order/cancel-all"
    url = base_url + path

    timestamp = str(int(time.time() * 1000))
    method = "POST"

    body_dict = {"symbol": symbol, "category": category}
    body_json = json.dumps(body_dict, separators=(",", ":"))

    signature = generate_signature(api_secret, timestamp, method, path, body_json)

    headers = {
        "Content-Type": "application/json",
        "X-BAPI-API-KEY": api_key,
        "X-BAPI-TIMESTAMP": timestamp,
        "X-BAPI-RECV-WINDOW": "5000",
        "X-BAPI-SIGN": signature,
    }

    print(f"[DEBUG] Request URL: {url}")
    print(f"[DEBUG] Request Headers: {headers}")
    print(f"[DEBUG] Request Body: {body_json}")

    response = requests.post(url, headers=headers, data=body_json)
    print(f"[DEBUG] Response Status Code: {response.status_code}")

    print(f"\n=== {symbol}（{category}） キャンセルレスポンス ===")
    print(response.text)
    return response.json()


def read_cancel_list_from_excel(
    file_path, sheet_name, symbol_col="A", category_col="B"
):
    wb = load_workbook(file_path, data_only=True)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"シート名 '{sheet_name}' がExcelファイルに存在しません。")
    ws = wb[sheet_name]
    cancel_list = []

    for row in ws.iter_rows(min_row=2):  # 1行目はヘッダー想定
        symbol = row[ord(symbol_col.upper()) - ord("A")].value
        category = row[ord(category_col.upper()) - ord("A")].value
        if symbol and category:
            cancel_list.append((symbol.strip(), category.strip()))
    return cancel_list


def main():
    config = load_config()  # 一つ上の階層のyaml読み込み想定
    api_key = config["bybit_demo"]["key"]
    api_secret = config["bybit_demo"]["secret"]

    excel_path = "./cancel_list.xlsx"
    sheet_name = "Orders"
    symbol_col = "A"
    category_col = "B"

    cancel_entries = read_cancel_list_from_excel(
        excel_path, sheet_name, symbol_col, category_col
    )

    for symbol, category in cancel_entries:
        cancel_all_orders(api_key, api_secret, symbol, category)

    resp = requests.get("https://api-testnet.bybit.com/v2/public/time")
    print(resp.json())


if __name__ == "__main__":
    main()
