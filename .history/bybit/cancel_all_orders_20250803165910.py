import json
import time
import hmac
import hashlib
import requests
from openpyxl import load_workbook
from load_config import load_config  # APIキー等をyamlなどから読み込む想定


def cancel_all_orders(api_key, api_secret, symbol, category):
    url = "https://api.bybit.com/v5/order/cancel-all"
    timestamp = str(int(time.time() * 1000))
    recv_window = "5000"

    body = {"symbol": symbol, "category": category}
    body_json = json.dumps(body, separators=(",", ":"))

    sign_payload = f"{timestamp}{api_key}{recv_window}{body_json}"
    signature = hmac.new(
        api_secret.encode(), sign_payload.encode(), hashlib.sha256
    ).hexdigest()

    headers = {
        "Content-Type": "application/json",
        "X-BAPI-API-KEY": api_key,
        "X-BAPI-TIMESTAMP": timestamp,
        "X-BAPI-RECV-WINDOW": recv_window,
        "X-BAPI-SIGN": signature,
    }

    response = requests.post(url, headers=headers, data=body_json)
    print(f"\n=== {symbol}（{category}） キャンセルレスポンス ===")
    print(response.text)
    return response.json()


def read_cancel_list_from_excel(
    file_path, sheet_name, symbol_col="A", category_col="B"
):
    wb = load_workbook(file_path, data_only=True)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"シート名 '{sheet_name}' がExcelファイルに存在しません。")
    ws = wb[sheet_name]
    cancel_list = []

    for row in ws.iter_rows(min_row=2):  # 1行目はヘッダーと仮定
        symbol = row[ord(symbol_col.upper()) - ord("A")].value
        category = row[ord(category_col.upper()) - ord("A")].value

        if symbol and category:
            cancel_list.append((symbol.strip(), category.strip()))
    return cancel_list


def main():
    config = load_config()  # ← 一つ上の階層のファイルを指定
    api_key = config["bybit_demo"]["key"]
    api_secret = config["bybit_demo"]["secret"]

    print("API Key:", api_key)
    print("API Secret:", api_secret)

    # Excelファイル設定
    excel_path = "./cancel_list.xlsx"
    sheet_name = "Orders"  # 実際のシート名に合わせて変更
    symbol_col = "A"  # 銘柄列
    category_col = "B"  # カテゴリ列（例: linear, inverse, option）

    cancel_entries = read_cancel_list_from_excel(
        excel_path, sheet_name, symbol_col, category_col
    )

    for symbol, category in cancel_entries:
        cancel_all_orders(api_key, api_secret, symbol, category)


if __name__ == "__main__":
    main()
