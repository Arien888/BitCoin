import requests
import time
import hmac
import hashlib
import base64
import yaml
import csv
from openpyxl import load_workbook

import os

# 自分のファイル（main.py）があるディレクトリを取得
BASE_DIR = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))


def load_config(path=os.path.join(BASE_DIR, "config.yaml")):
    with open(path, "r", encoding="utf-8") as f:
        return yaml.safe_load(f)


def get_assets(api_key, api_secret, api_passphrase, request_path, product_type=None):
    print(f"APIリクエスト: {request_path} (product_type={product_type})")

    method = "GET"
    timestamp = str(int(time.time() * 1000))
    # クエリパラメータがある場合は request_path に追加し、署名対象文字列にも含める
    if product_type:
        query_string = f"?productType={product_type}"
        full_path = request_path + query_string
    else:
        query_string = ""
        full_path = request_path

    prehash_string = timestamp + method + full_path + ""

    signature = hmac.new(
        api_secret.encode("utf-8"), prehash_string.encode("utf-8"), hashlib.sha256
    ).digest()
    signature_base64 = base64.b64encode(signature).decode()

    headers = {
        "ACCESS-KEY": api_key,
        "ACCESS-SIGN": signature_base64,
        "ACCESS-TIMESTAMP": timestamp,
        "ACCESS-PASSPHRASE": api_passphrase,
        "Content-Type": "application/json",
        "locale": "en-US",
    }

    url = base_url.rstrip('/') + full_path
    print(f"APIリクエスト full_path: {full_path}")
    response = requests.get(url, headers=headers)
    response.raise_for_status()
    print(f"APIリクエスト成功: {url}")
    return response.json()


def save_assets_to_csv_jp(filename, data, keys):
    if data is None or not isinstance(data, dict):
        print(f"{filename}: データが不正です")
        return

    assets = data.get("data", [])
    if not assets:
        print(f"{filename}: 資産データがありません")
        return

    with open(filename, mode="w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(keys)
        for asset in assets:
            row = [asset.get(k, "") for k in keys]
            writer.writerow(row)

    print(f"{filename} にCSV出力しました")


def write_to_existing_excel(filename, data, keys, sheet_name="Sheet1"):
    """
    既存のExcelファイルにdataを書き込む（上書きまたは追記）
    - filename: Excelファイル名（既存ファイル）
    - data: 辞書リストまたは dict の 'data' 部分。例: data["data"]
    - keys: 書き込みたいキーのリスト（列ヘッダーにする）
    - sheet_name: 書き込み先シート名（デフォルト"Sheet1"）

    既存ファイルの先頭にヘッダーを書き、続けてデータを書き込む（シートは上書きされる）
    """
    # ファイル読み込み
    wb = load_workbook(filename)

    # シート取得 or 新規作成
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        # 既存シートの内容をクリア（全データ削除）
        ws.delete_rows(1, ws.max_row)
    else:
        ws = wb.create_sheet(sheet_name)

    # ヘッダー書き込み
    ws.append(keys)

    # データ書き込み
    assets = data.get("data", [])
    for asset in assets:
        row = [asset.get(k, "") for k in keys]
        ws.append(row)

    # 保存
    wb.save(filename)
    print(f"{filename} に既存ファイル上書き保存しました")
