import requests
import time
import hmac
import hashlib
import base64
import yaml
import csv
from openpyxl import load_workbook
from urllib.parse import urlparse  # ← 必要
import os

# 自分のファイル（main.py）があるディレクトリを取得
BASE_DIR = os.path.dirname(os.path.abspath(__file__))


def load_config(path=os.path.join(BASE_DIR, "config.yaml")):
    with open(path, "r") as f:
        return yaml.safe_load(f)


def get_futures_eccout_equity(
    api_key, api_secret, api_passphrase, base_url="https://api.bitget.com"
):
    timestamp = str(int(time.time() * 1000))
    method = "GET"
    request_path = "/api/v2/mix/account/account"
    full_path = request_path  # クエリなし

    pre_hash = timestamp + method + full_path
    signature = hmac.new(
        api_secret.encode("utf-8"), pre_hash.encode("utf-8"), hashlib.sha256
    ).digest()
    signature_base64 = base64.b64encode(signature).decode()

    headers = {
        "ACCESS-KEY": api_key,
        "ACCESS-SIGN": signature_base64,
        "ACCESS-TIMESTAMP": timestamp,
        "ACCESS-PASSPHRASE": api_passphrase,
        "Content-Type": "application/json",
    }
    url = base_url + full_path
    response = requests.get(url, headers=headers)
    response.raise_for_status()
    return response.json()

def get_assets(api_key, api_secret, api_passphrase, request_path, product_type=None, margin_coin=None):
    import time, hmac, hashlib, base64, requests

    base_url = "https://api.bitget.com"
    method = "GET"
    timestamp = str(int(time.time() * 1000))

    query_string = ""
    if product_type:
        query_string = f"?productType={product_type}"
    # 他のパラメータがあれば同様に付加する

    full_path = request_path + query_string
    body = ""

    prehash_string = timestamp + method + full_path + body
    signature = hmac.new(api_secret.encode("utf-8"), prehash_string.encode("utf-8"), hashlib.sha256).digest()
    signature_base64 = base64.b64encode(signature).decode()

    headers = {
        "ACCESS-KEY": api_key,
        "ACCESS-SIGN": signature_base64,
        "ACCESS-TIMESTAMP": timestamp,
        "ACCESS-PASSPHRASE": api_passphrase,
        "Content-Type": "application/json",
        "locale": "en-US",
    }

    url = base_url + full_path

    response = requests.get(url, headers=headers)

    if response.status_code != 200:
        print("HTTP Error:", response.status_code)
        return None

    data = response.json()
    if data.get("code") != "00000":
        print("API Error:", data.get("msg"))
        return None

    return data


def save_assets_to_csv_jp(filename, data, keys):
    if data is None or not isinstance(data, dict):
        print(f"{filename}: データが不正です")
        return

    assets = data.get("data", [])
    if not assets:
        print(f"{filename}: 資産データがありません")
        return

    with open(filename, mode="w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(keys)
        for asset in assets:
            row = [asset.get(k, "") for k in keys]
            writer.writerow(row)

    print(f"{filename} にCSV出力しました")


def write_to_existing_excel(filename, data, keys, sheet_name="Sheet1"):
    """
    既存のExcelファイルにdataを書き込む（上書きまたは追記）
    - filename: Excelファイル名（既存ファイル）
    - data: 辞書リストまたは dict の 'data' 部分。例: data["data"]
    - keys: 書き込みたいキーのリスト（列ヘッダーにする）
    - sheet_name: 書き込み先シート名（デフォルト"Sheet1"）

    既存ファイルの先頭にヘッダーを書き、続けてデータを書き込む（シートは上書きされる）
    """
    # ファイル読み込み
    wb = load_workbook(filename)

    # シート取得 or 新規作成
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        # 既存シートの内容をクリア（全データ削除）
        ws.delete_rows(1, ws.max_row)
    else:
        ws = wb.create_sheet(sheet_name)

    # ヘッダー書き込み
    ws.append(keys)

    # データ書き込み
    assets = data.get("data", [])
    for asset in assets:
        row = [asset.get(k, "") for k in keys]
        ws.append(row)

    # 保存
    wb.save(filename)
    print(f"{filename} に既存ファイル上書き保存しました")

def get_futures_account(api_key, api_secret, api_passphrase, product_type="UMCBL"):
    import time, hmac, hashlib, base64, requests

    valid_product_types = {"UMCBL", "CMCBL", "UMCSP", "CMCSL"}
    if product_type not in valid_product_types:
        raise ValueError(f"Invalid productType: {product_type}. Choose from {valid_product_types}")

    base_url = "https://api.bitget.com"
    method = "GET"
    request_path = "/api/v2/mix/account/accounts"
    timestamp = str(int(time.time() * 1000))

    query_string = f"?productType={product_type}"
    full_path = request_path + query_string
    body = ""

    prehash_string = timestamp + method + full_path + body
    signature = hmac.new(api_secret.encode("utf-8"), prehash_string.encode("utf-8"), hashlib.sha256).digest()
    signature_base64 = base64.b64encode(signature).decode()

    headers = {
        "ACCESS-KEY": api_key,
        "ACCESS-SIGN": signature_base64,
        "ACCESS-TIMESTAMP": timestamp,
        "ACCESS-PASSPHRASE": api_passphrase,
        "Content-Type": "application/json",
        "locale": "en-US",
    }

    url = base_url + full_path
    response = requests.get(url, headers=headers)

    if response.status_code != 200:
        print(f"HTTP Error: {response.status_code} - {response.text}")
        return None

    data = response.json()
    if data.get("code") != "00000":
        print(f"API Error: {data.get('msg')}")
        return None

    return data
