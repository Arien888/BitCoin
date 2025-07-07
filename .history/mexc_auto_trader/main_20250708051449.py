import openpyxl
import requests
import yaml
import time
import hmac
import hashlib
import os

# === 設定ファイル読み込み ===
CONFIG_PATH = os.path.join(os.path.dirname(__file__), "..", "config.yaml")
with open(CONFIG_PATH, "r") as f:
    config = yaml.safe_load(f)

API_KEY = config["mexc"]["api_key"]
API_SECRET = config["mexc"]["api_secret"]

# === Excelファイルパス ===
EXCEL_PATH = os.path.join(os.path.dirname(__file__), "orders.xlsx")


# === 署名作成 ===
def generate_signature(api_key, api_secret, req_time, sign_params):
    sign_string = f"{api_key}{req_time}{sign_params}"
    return hmac.new(
        api_secret.encode(), sign_string.encode(), hashlib.sha256
    ).hexdigest()


# === 発注関数 ===
def place_order(symbol, side, price, quantity, order_type):
    url = "https://contract.mexc.com/api/v1/private/order/submit"

    req_time = str(int(time.time() * 1000))

    params = {
        "symbol": symbol,
        "price": price,
        "vol": quantity,
        "side": 1 if side == "BUY" else 2,
        "type": 1 if order_type == "LIMIT" else 2,
        "open_type": 1,
        "position_id": 0,
        "leverage": 20,
        "external_oid": "",
        "stop_loss_price": 0,
        "take_profit_price": 0,
        "position_mode": 1,
    }

    # 署名用にパラメータを文字列化
    param_str = "&".join(f"{k}={v}" for k, v in sorted(params.items()))
    signature = generate_signature(API_KEY, API_SECRET, req_time, param_str)

    headers = {
        "Content-Type": "application/json",
        "ApiKey": API_KEY,
        "Request-Time": req_time,
        "Signature": signature,
    }

    print("→ MEXCに発注送信中...")
    try:
        response = requests.post(
            url, json=params, headers=headers, timeout=15, verify=True
        )

        print("← レスポンス受信:", response.json())
        return response.json()
    except Exception as e:
        print("❌ 発注エラー:", e)
        return {"error": str(e)}


# === ログ保存 ===
def log_result(row, result):
    log_path = os.path.join(os.path.dirname(__file__), "order_log.txt")
    with open(log_path, "a", encoding="utf-8") as f:
        f.write(f"{time.strftime('%Y-%m-%d %H:%M:%S')} - {row} → {result}\n")


# === メイン関数 ===
def main():
    wb = openpyxl.load_workbook(EXCEL_PATH)
    sheet = wb.active

    print("📋 Excel注文データ一覧:")
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if all(cell is None for cell in row):
            continue  # 空行スキップ
        print(row)

    for row in sheet.iter_rows(min_row=2, values_only=True):
        if all(cell is None for cell in row):
            continue
        symbol, side, price, quantity, order_type = row
        print(f"▶️ 発注中: {symbol}, {side}, {price}, {quantity}, {order_type}")
        result = place_order(symbol, side, price, quantity, order_type)
        log_result(row, result)
        time.sleep(1)


# === 実行 ===
if __name__ == "__main__":
    main()
