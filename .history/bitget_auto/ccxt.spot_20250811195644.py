import os
import yaml
import ccxt
from openpyxl import load_workbook

# 設定ファイルのパス
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
CONFIG_PATH = os.path.abspath(os.path.join(BASE_DIR, "..", "config.yaml"))

# config.yaml 読み込み
with open(CONFIG_PATH, "r", encoding="utf-8") as f:
    config = yaml.safe_load(f)

# APIキー等
api_key = config["bitget"]["api_key"]
secret = config["bitget"]["api_secret"]
password = config["bitget"]["passphrase"]  # Bitgetはパスワード必須

excel_rel_path = config["order_export"]["source_file"]

# Excelファイルの絶対パス
excel_path = os.path.abspath(os.path.join(BASE_DIR, "..", excel_rel_path))

# 注文を読み込むシート名
sheet_names = [
    config["excel"]["sheets"]["bitget_open_long_spot"],
    config["excel"]["sheets"]["bitget_close_long_spot"],
    # 他シートあれば追加
]

# Bitget クライアント作成
bitget = ccxt.bitget({
    "apiKey": api_key,
    "secret": secret,
    "password": password,  # Bitget固有
})

# Excel読み込み
wb = load_workbook(excel_path, data_only=True)

for sheet_name in sheet_names:
    if sheet_name not in wb.sheetnames:
        print(f"シート {sheet_name} がExcelファイルに存在しません")
        continue

    ws = wb[sheet_name]

    # 1行目はヘッダーなので2行目から開始
    for row in ws.iter_rows(min_row=2, values_only=True):
        symbol, side, order_type, size, price = row[:5]

        if not all([symbol, side, order_type, size]):
            print(f"不足データのためスキップ: {row}")
            continue

        # Bitget現物のシンボル形式は例: BTCUSDT_SPBL
        if not symbol.endswith("_SPBL"):
            symbol = symbol.upper() + "_SPBL"

        # 注文タイプ変換
        order_type_ccxt = order_type.lower()
        if order_type_ccxt not in ["limit", "market"]:
            print(f"対応していない注文タイプ: {order_type}（スキップ）")
            continue

        # side変換
        side_ccxt = side.lower()
        if side_ccxt not in ["buy", "sell"]:
            print(f"対応していない注文方向: {side}（スキップ）")
            continue

        try:
            if order_type_ccxt == "limit":
                order = bitget.create_limit_order(symbol, side_ccxt, size, price)
            elif order_type_ccxt == "market":
                order = bitget.create_market_order(symbol, side_ccxt, size)

            print(f"注文成功: {order}")
        except Exception as e:
            print(
                f"注文エラー: symbol={symbol}, side={side}, type={order_type}, size={size}, price={price}"
            )
            print("エラー内容:", e)
