# main.py

import sys
import io
import os
import yaml
import openpyxl

from bitget_client import BitgetClient

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")

CONFIG_PATH = os.path.join(os.path.dirname(__file__), "..", "config.yaml")

with open(CONFIG_PATH, "r", encoding="utf-8") as f:
    config = yaml.safe_load(f)

excel_path = config["excel"]["path"]
client = BitgetClient(
    key=config["bitget"]["api_key"],
    secret=config["bitget"]["api_secret"],
    passphrase=config["bitget"]["passphrase"],
)


def read_orders_from_excel(file_path, sheet_name):
    wb = openpyxl.load_workbook(file_path)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"シート '{sheet_name}' が見つかりません。")

    ws = wb[sheet_name]

    orders = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        symbol, side, price, quantity, order_type = row
        orders.append((symbol, side, price, quantity, order_type))
    return orders


def main():
    buy_sheet = config["excel"]["sheets"]["buy"]
    sell_sheet = config["excel"]["sheets"]["sell"]

    print("[INFO] Buyシートから読み込み:")
    buy_orders = read_orders_from_excel(excel_path, buy_sheet)
    for order in buy_orders:
        print(order)
        client.place_order(*order)

    print("[INFO] Sellシートから読み込み:")
    sell_orders = read_orders_from_excel(excel_path, sell_sheet)
    for order in sell_orders:
        print(order)
        client.place_order(*order)


if __name__ == "__main__":
    main()
