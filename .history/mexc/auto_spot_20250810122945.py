import os
import yaml
import ccxt
from openpyxl import load_workbook

# 設定ファイルのパスを組み立て
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
CONFIG_PATH = os.path.abspath(os.path.join(BASE_DIR, "..", "config.yaml"))

# config.yaml 読み込み
with open(CONFIG_PATH, "r", encoding="utf-8") as f:
    config = yaml.safe_load(f)

api_key = config["mexc"]["api_key"]
secret = config["mexc"]["api_secret"]
excel_rel_path = config["order_export"]["source_file"]

# Excelファイルの絶対パス
excel_path = os.path.abspath(os.path.join(BASE_DIR, "..", excel_rel_path))

# 注文を読み込むシート名（例では一つ目）
sheet_names = [
    config["excel"]["sheets"]["mexc_open_long_spot"],
    # 他のシートがあれば追加
]

# MEXC クライアント作成
mexc = ccxt.mexc({"apiKey": api_key, "secret": secret})

# Excel読み込み
wb = load_workbook(excel_path, data_only=True)

for sheet_name in sheet_names:
    if sheet_name not in wb.sheetnames:
        print(f"シート {sheet_name} がExcelファイルに存在しません")
        continue

    ws = wb[sheet_name]

    # 1行目はヘッダーなので2行目から開始
    for row in ws.iter_rows(min_row=2, values_only=True):
        symbol, side, order_type, size, price = row[:5]

        if not all([symbol, side, order_type, size]):
            # 必要な情報が無ければスキップ
            print(f"不足データのためスキップ: {row}")
            continue

        # ccxtの注文タイプ変換（order_typeがリミットなら'limit'、成行なら'market'等）
        order_type_ccxt = order_type.lower()
        if order_type_ccxt not in ["limit", "market"]:
            print(f"対応していない注文タイプ: {order_type}（スキップ）")
            continue

        # sideはbuy/sellに変換（大文字小文字混在対策）
        side_ccxt = side.lower()
        if side_ccxt not in ["buy", "sell"]:
            print(f"対応していない注文方向: {side}（スキップ）")
            continue

        try:
            if order_type_ccxt == "limit":
                # price指定あり
                order = mexc.create_limit_order(symbol, side_ccxt, size, price)
            elif order_type_ccxt == "market":
                # 成行ならprice不要
                order = mexc.create_market_order(symbol, side_ccxt, size)
            print(f"注文成功: {order}")
        except Exception as e:
            print(
                f"注文エラー: symbol={symbol}, side={side}, type={order_type}, size={size}, price={price}"
            )
            print("エラー内容:", e)
