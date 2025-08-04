import yaml
from pathlib import Path
from openpyxl import load_workbook, Workbook
import time
from openpyxl.utils import get_column_letter  # ãƒ•ã‚¡ã‚¤ãƒ«å†’é ­ã® import ã«è¿½åŠ 

# âœ… YAMLã®èª­ã¿è¾¼ã¿
config_path = Path(__file__).resolve().parent.parent / "config.yaml"
with open(config_path, "r", encoding="utf-8") as f:
    config = yaml.safe_load(f)

# âœ… è¨­å®šå–å¾—
order_cfg = config["order_export"]
src_file = (config_path.parent / order_cfg["source_file"]).resolve()
dst_file = (config_path.parent / order_cfg["output_file"]).resolve()

# âœ… ã‚·ãƒ¼ãƒˆåã®å–å¾—ï¼ˆsheet_name ã‚‚è¨±å®¹ï¼‰
sheet_names = order_cfg.get("sheet_names") or [order_cfg.get("sheet_name")]
if not sheet_names or None in sheet_names:
    raise KeyError(
        "YAMLã« 'sheet_name' ã¾ãŸã¯ 'sheet_names' ãŒæ­£ã—ãè¨­å®šã•ã‚Œã¦ã„ã¾ã›ã‚“ã€‚"
    )


def export_values_only(src_file, sheet_names, dst_file):
    print(f"ğŸ“„ èª­ã¿è¾¼ã¿å…ƒ: {src_file}")
    print(f"ğŸ“¤ å‡ºåŠ›å…ˆ: {dst_file}")

    time.sleep(1.5)  # â† ãƒ•ã‚¡ã‚¤ãƒ«é–‹ãå‰ã«ä¸€å‘¼å¸

    src_wb = load_workbook(src_file, data_only=True)
    new_wb = Workbook()
    first = True

    for name in sheet_names:
        if name not in src_wb.sheetnames:
            print(f"âš  ã‚·ãƒ¼ãƒˆ '{name}' ã¯å­˜åœ¨ã—ã¾ã›ã‚“ã€‚ã‚¹ã‚­ãƒƒãƒ—ã—ã¾ã™ã€‚")
            continue

        print(f"âœ… ã‚·ãƒ¼ãƒˆå‡¦ç†ä¸­: {name}")
        src_ws = src_wb[name]
        if first:
            new_ws = new_wb.active
            new_ws.title = name
            first = False
        else:
            new_ws = new_wb.create_sheet(title=name)

        for row in src_ws.iter_rows():
            for cell in row:
                v = cell.value

                # âœ… æ–‡å­—åˆ—ã¨ã—ã¦æ›¸ã‹ã‚ŒãŸæ•°å€¤ã‚’ float ã«å¤‰æ›
                if isinstance(v, str):
                    try:
                        v = float(v)
                    except ValueError:
                        pass  # æ•°å€¤ã§ãªã‘ã‚Œã°ãã®ã¾ã¾

                new_ws[cell.coordinate].value = v

    print("â³ ä¿å­˜æº–å‚™ä¸­...")
    time.sleep(1.0)

    new_wb.save(dst_file)
    print(f"âœ… ã‚¨ã‚¯ã‚¹ãƒãƒ¼ãƒˆå®Œäº†: {dst_file}")


# âœ… å®Ÿè¡Œ
if __name__ == "__main__":
    export_values_only(src_file, sheet_names, dst_file)
