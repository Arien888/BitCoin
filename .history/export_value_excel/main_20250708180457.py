from openpyxl import load_workbook, Workbook

def export_values_only(src_file, sheet_name, dst_file):
    wb = load_workbook(src_file, data_only=True)
    ws = wb[sheet_name]

    new_wb = Workbook()
    new_ws = new_wb.active

    for row in ws.iter_rows():
        for cell in row:
            new_ws[cell.coordinate].value = cell.value

    new_wb.save(dst_file)

# 使い方
export_values_only("元ファイル.xlsx", "Sheet1", "値のみファイル.xlsx")
