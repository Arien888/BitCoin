import yaml
from pathlib import Path
from openpyxl import load_workbook, Workbook
import time

# ✅ YAMLの読み込み
config_path = Path(__file__).resolve().parent.parent / "config.yaml"
with open(config_path, "r", encoding="utf-8") as f:
    config = yaml.safe_load(f)

# ✅ 設定取得
order_cfg = config["order_export"]
src_file = (config_path.parent / order_cfg["source_file"]).resolve()
dst_file = (config_path.parent / order_cfg["output_file"]).resolve()

# ✅ シート名の取得（sheet_name も許容）
sheet_names = order_cfg.get("sheet_names") or [order_cfg.get("sheet_name")]
if not sheet_names or None in sheet_names:
    raise KeyError(
        "YAMLに 'sheet_name' または 'sheet_names' が正しく設定されていません。"
    )


# ✅ 値をコピーしつつ、文字列で書かれた数値を float に変換して書き込む関数
def export_values_only(src_file, sheet_names, dst_file):
    print(f"📄 読み込み元: {src_file}")
    print(f"📤 出力先: {dst_file}")

    time.sleep(1.5)  # ← ファイル開く前に一呼吸

    src_wb = load_workbook(src_file, data_only=True)
    new_wb = Workbook()
    first = True

    for name in sheet_names:
        if name not in src_wb.sheetnames:
            print(f"⚠ シート '{name}' は存在しません。スキップします。")
            continue

        print(f"✅ シート処理中: {name}")
        src_ws = src_wb[name]
        if first:
            new_ws = new_wb.active
            new_ws.title = name
            first = False
        else:
            new_ws = new_wb.create_sheet(title=name)

        for row in src_ws.iter_rows():
            for cell in row:
                v = cell.value
                if v is None:
                    v = str(v)  # None を空文字列に変換
                new_ws[cell.coordinate].value = v

    print("⏳ 保存準備中...")
    time.sleep(1.0)  # ← 保存前に安定化のため少し待つ

    new_wb.save(dst_file)
    print(f"✅ エクスポート完了: {dst_file}")


# ✅ 実行
if __name__ == "__main__":
    export_values_only(src_file, sheet_names, dst_file)
