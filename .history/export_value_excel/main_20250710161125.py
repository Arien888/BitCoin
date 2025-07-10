import yaml
from pathlib import Path
from openpyxl import load_workbook, Workbook

# ✅ YAMLの読み込み
config_path = Path(__file__).resolve().parent / "config.yaml"
with open(config_path, "r", encoding="utf-8") as f:
    config = yaml.safe_load(f)

# ✅ 設定取得
order_cfg = config["order_export"]
src_file = (config_path.parent / order_cfg["source_file"]).resolve()
dst_file = (config_path.parent / order_cfg["output_file"]).resolve()

# ✅ シート名取得
sheet_names = order_cfg.get("sheet_names") or [order_cfg.get("sheet_name")]
if not sheet_names or None in sheet_names:
    raise KeyError("YAMLに 'sheet_name' または 'sheet_names' が正しく設定されていません。")

# ✅ 値だけを数値型でエクスポートする関数
def export_values_only(src_file, sheet_names, dst_file):
    print(f"📄 読み込み元: {src_file}")
    print(f"📤 出力先: {dst_file}")

    src_wb = load_workbook(src_file, data_only=True)
    new_wb = Workbook()
    first = True

    for name in sheet_names:
        if name not in src_wb.sheetnames:
            print(f"⚠ シート '{name}' は存在しません。スキップします。")
            continue

        print(f"✅ シート処理中: {name}")
        src_ws = src_wb[name]
        if first:
            new_ws = new_wb.active
            new_ws.title = name
            first = False
        else:
            new_ws = new_wb.create_sheet(title=name)

        for row in src_ws.iter_rows():
            for cell in row:
                v = cell.value
                # ✅ 数値変換を試みる（文字列→float）
                if isinstance(v, str):
                    try:
                        v = float(v)
                    except ValueError:
                        pass  # 数値でなければそのまま
                new_ws[cell.coordinate].value = v

    new_wb.save(dst_file)
    print(f"✅ エクスポート完了: {dst_file}")

# ✅ 実行
if __name__ == "__main__":
    export_values_only(src_file, sheet_names, dst_file)
