import requests
import datetime as dt
import yaml
import os
import pandas as pd
from openpyxl import load_workbook

symbols = [
    "BTC/JPY",
    "XRP/JPY",
    "LTC/JPY",
    "ETH/JPY",
    "MONA/JPY",
    "BCC/JPY",
    "XLM/JPY",
    "QTUM/JPY",
    "BAT/JPY",
    "OMG/JPY",
    "XYM/JPY",
    "LINK/JPY",
    "MKR/JPY",
    "BOBA/JPY",
    "ENJ/JPY",
    "POL/JPY",
    "DOT/JPY",
    "DOGE/JPY",
    "ASTR/JPY",
    "ADA/JPY",
    "AVAX/JPY",
    "AXS/JPY",
    "FLR/JPY",
    "SAND/JPY",
    "APE/JPY",
    "GALA/JPY",
    "CHZ/JPY",
    "OAS/JPY",
    "MANA/JPY",
    "GRT/JPY",
    "RENDER/JPY",
    "BNB/JPY",
    "ARB/JPY",
    "OP/JPY",
    "DAI/JPY",
    "KLAY/JPY",
    "IMX/JPY",
    "MASK/JPY",
    "SOL/JPY",
    "CYBER/JPY",
    "TRX/JPY",
    "LPT/JPY",
    "ATOM/JPY",
]


def load_config():
    # main.py の1つ上の階層にconfig.yamlがある想定
    base_dir = os.path.dirname(os.path.abspath(__file__))
    config_path = os.path.join(base_dir, "..", "config.yaml")
    with open(config_path, "r", encoding="utf-8") as f:
        config = yaml.safe_load(f)
    return config


def fetch_prices(symbols):
    prices = []
    for symbol in symbols:
        url = f"https://public.bitbank.cc/{symbol}/ticker"
        try:
            response = requests.get(url)
            res_json = response.json()
            if res_json.get("success", False):
                last_price = float(res_json["data"]["last"])
                prices.append((symbol.upper(), last_price))
            else:
                print(f"APIエラー: {symbol} - {res_json}")
        except Exception as e:
            print(f"取得失敗: {symbol} - {e}")
    return prices


def write_prices_to_excel(prices, file_path, sheet_name="bitbank_now_price"):
    df_prices = pd.DataFrame(prices, columns=["Symbol", "LastPrice"])

    if not os.path.exists(file_path):
        # ファイルがなければ新規作成
        with pd.ExcelWriter(file_path, engine="openpyxl") as writer:
            df_prices.to_excel(writer, sheet_name=sheet_name, index=False)
    else:
        # ファイルがあれば既存シートがあれば削除してから書き込み
        book = load_workbook(file_path)
        if sheet_name in book.sheetnames:
            std = book[sheet_name]
            book.remove(std)
            book.save(file_path)
        book.close()

        with pd.ExcelWriter(
            file_path, engine="openpyxl", mode="a", if_sheet_exists="new"
        ) as writer:
            df_prices.to_excel(writer, sheet_name=sheet_name, index=False)


def main():
    config = load_config()
    excel_path = config["order_export"]["writer_file"]

    now = dt.datetime.now()
    prices = fetch_prices(symbols)

    print(f"=== {now} 現在価格 ===")
    for sym, price in prices:
        print(f"{sym}: {price}")

    write_prices_to_excel(prices, excel_path)


if __name__ == "__main__":
    main()
