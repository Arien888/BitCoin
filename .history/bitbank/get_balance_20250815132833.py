import os
import yaml
import ccxt
from openpyxl import load_workbook
from datetime import datetime

# --- 設定ファイルの読み込み ---
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
CONFIG_PATH = os.path.abspath(os.path.join(BASE_DIR, "..", "config.yaml"))

with open(CONFIG_PATH, "r", encoding="utf-8") as f:
    config = yaml.safe_load(f)

api_key = config["bitbank"]["api_key"]
secret = config["bitbank"]["api_secret"]

# --- Excelファイルのパス ---
excel_rel_path = config["order_export"]["source_file"]
excel_path = os.path.abspath(os.path.join(BASE_DIR, "..", excel_rel_path))

# --- Bitbank クライアント作成 ---
bitbank = ccxt.bitbank({
    "apiKey": api_key,
    "secret": secret,
})

# --- 残高取得 ---
try:
    balance = bitbank.fetch_balance()
except Exception as e:
    print("残高取得エラー:", e)
    balance = {}

# --- Excel読み込み ---
wb = load_workbook(excel_path)

# --- 新規シート名（日付入りで上書き防止） ---
now_str = datetime.now().strftime("%Y%m%d_%H%M%S")
sheet_name = f"bitbank_balance_{now_str}"
ws = wb.create_sheet(sheet_name)

# ヘッダー
ws.append(["銘柄", "数量"])

# 保有銘柄と数量を書き込む
for asset, amount in balance.get('total', {}).items():
    if amount and amount > 0:
        ws.append([asset, amount])

# Excel保存
wb.save(excel_path)
print(f"Excelの新規シート '{sheet_name}' に残高を書き込みました。")
