import os
import yaml
import ccxt
from openpyxl import load_workbook

# --- 設定ファイルの読み込み ---
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
CONFIG_PATH = os.path.abspath(os.path.join(BASE_DIR, "..", "config.yaml"))

with open(CONFIG_PATH, "r", encoding="utf-8") as f:
    config = yaml.safe_load(f)

api_key = config["bitbank"]["api_key"]
secret = config["bitbank"]["api_secret"]

# --- Excelファイルのパス ---
excel_rel_path = config["order_export"]["source_file"]
excel_path = os.path.abspath(os.path.join(BASE_DIR, "..", excel_rel_path))

# --- Bitbank クライアント作成 ---
bitbank = ccxt.bitbank(
    {
        "apiKey": api_key,
        "secret": secret,
    }
)

# --- 残高取得 ---
try:
    balance = bitbank.fetch_balance()
except Exception as e:
    print("残高取得エラー:", e)
    balance = {}

# --- Excel読み込み ---
wb = load_workbook(excel_path)


# --- bitbank_balance シート作成または取得 ---
sheet_name = "bitbank_balance"
if sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    # 既存データをクリア
    ws.delete_rows(1, ws.max_row)
else:
    ws = wb.create_sheet(sheet_name)

# ヘッダーを書き込む
ws.append(["銘柄", "数量"])

# 保有銘柄と数量を書き込む
for asset, amount in balance.get("total", {}).items():
    if amount and amount > 0:
        ws.append([asset, amount])

# Excel保存
wb.save(excel_path)
print(f"Excelのシート '{sheet_name}' に残高を書き込みました。")

# --- 有効残高を表示 ---
if balance:
    print("保有銘柄と数量:")
    # balance['total'] に各通貨の総保有量が入っている
    for asset, amount in balance.get("total", {}).items():
        if amount and amount > 0:
            print(f"{asset}: {amount}")
