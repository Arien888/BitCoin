import yaml
from pathlib import Path
from openpyxl import load_workbook, Workbook

# YAMLの読み込み
config_path = Path(__file__).resolve().parent.parent / "config.yaml"
with open(config_path, "r", encoding="utf-8") as f:
    config = yaml.safe_load(f)

# 設定取得（パスはプロジェクトルート基準で補正）
order_cfg = config["order_export"]
src_file = (config_path.parent / order_cfg["source_file"]).resolve()
dst_file = (config_path.parent / order_cfg["output_file"]).resolve()
sheet_name = order_cfg["sheet_name"]

# エクスポート関数
def export_values_only(src_file, sheet_name, dst_file):
    wb = load_workbook(src_file, data_only=True)
    ws = wb[sheet_name]

    new_wb = Workbook()
    new_ws = new_wb.active
    new_ws.title = sheet_name

    for row in ws.iter_rows():
        for cell in row:
            new_ws[cell.coordinate].value = cell.value

    new_wb.save(dst_file)

# ✅ 実行（YAMLから取得した情報で）
export_values_only(src_file, sheet_name, dst_file)
