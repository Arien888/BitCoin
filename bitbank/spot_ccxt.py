import os
import yaml
import ccxt
from openpyxl import load_workbook

# 設定ファイルのパスを組み立て
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
CONFIG_PATH = os.path.abspath(os.path.join(BASE_DIR, "..", "config.yaml"))

# config.yaml 読み込み
with open(CONFIG_PATH, "r", encoding="utf-8") as f:
    config = yaml.safe_load(f)

# Bitbank APIキーとシークレット取得
api_key = config["bitbank"]["api_key"]
secret = config["bitbank"]["api_secret"]

# Excelファイルのパス
excel_rel_path = config["order_export"]["source_file"]
excel_path = os.path.abspath(os.path.join(BASE_DIR, "..", excel_rel_path))

# 注文シート名
sheet_names = [
    config["excel"]["sheets"]["bitbank_open_long_spot"],
    # config["excel"]["sheets"]["bitbank_close_long_spot"],
]

# Bitbank クライアント作成
bitbank = ccxt.bitbank({
    "apiKey": api_key,
    "secret": secret,
})

# Excel読み込み
wb = load_workbook(excel_path, data_only=True)

for sheet_name in sheet_names:
    if sheet_name not in wb.sheetnames:
        print(f"シート {sheet_name} がExcelファイルに存在しません")
        continue

    ws = wb[sheet_name]

    # 1行目はヘッダーなので2行目から
    for row in ws.iter_rows(min_row=2, values_only=True):
        symbol, side, order_type, size, price = row[:5]

        if not all([symbol, side, order_type, size]):
            print(f"不足データのためスキップ: {row}")
            continue

        symbol_ccxt = symbol.strip()   # 空白削除のみ

        # 注文タイプ変換
        order_type_ccxt = order_type.lower()
        if order_type_ccxt not in ["limit", "market"]:
            print(f"対応していない注文タイプ: {order_type}（スキップ）")
            continue

        # 売買方向変換
        side_ccxt = side.lower()
        if side_ccxt not in ["buy", "sell"]:
            print(f"対応していない注文方向: {side}（スキップ）")
            continue

        try:
            if order_type_ccxt == "limit":
                order = bitbank.create_limit_order(symbol_ccxt, side_ccxt, size, price)
            elif order_type_ccxt == "market":
                order = bitbank.create_market_order(symbol_ccxt, side_ccxt, size)
            print(f"注文成功: {order}")
        except Exception as e:
            print(
                f"注文エラー: symbol={symbol_ccxt}, side={side}, type={order_type}, size={size}, price={price}"
            )
            print("エラー内容:", e)
