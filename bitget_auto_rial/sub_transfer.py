import hmac
import hashlib
import base64
import time
import requests
import uuid
import json
import os
import yaml
from openpyxl import load_workbook

# --- スクリプトのあるディレクトリ ---
BASE_DIR = os.path.dirname(os.path.abspath(__file__))

# --- config.yaml 読み込み ---
with open(os.path.join(BASE_DIR, "..", "config.yaml"), "r", encoding="utf-8") as f:
    config = yaml.safe_load(f)

API_KEY = config["bitget"]["api_key"]
API_SECRET = config["bitget"]["api_secret"]
PASSPHRASE = config["bitget"]["passphrase"]
IS_TESTNET = config["bitget"].get("is_testnet", False)
BASE_URL = "https://api.bitget.com" if not IS_TESTNET else "https://api-testnet.bitget.com"

MAIN_UID = str(config["bitget"]["main_uid"])
SUB_UID = str(config["bitget"]["subaccount"]["sub_uid"])

excel_path = os.path.join(BASE_DIR, "..", config["order_export"]["source_file"])
sheet_name = config["excel"]["sheets"]["bitget_sub_transfer"]

# ====== 署名作成 ======
def sign(method, path, body=""):
    timestamp = str(int(time.time() * 1000))
    message = timestamp + method + path + body
    mac = hmac.new(API_SECRET.encode("utf-8"), message.encode("utf-8"), hashlib.sha256)
    sign = base64.b64encode(mac.digest()).decode()
    return timestamp, sign

# ====== サブアカウント振替 ======
def transfer(from_uid, to_uid, from_type, to_type, coin, amount):
    path = "/api/v2/spot/wallet/subaccount-transfer"
    url = BASE_URL + path
    body_dict = {
        "fromUserId": str(from_uid),
        "toUserId": str(to_uid),
        "fromType": from_type,
        "toType": to_type,
        "coin": coin,
        "amount": str(amount),
        "clientOid": str(uuid.uuid4())
    }
    body = json.dumps(body_dict)
    timestamp, signature = sign("POST", path, body)
    headers = {
        "ACCESS-KEY": API_KEY,
        "ACCESS-SIGN": signature,
        "ACCESS-TIMESTAMP": timestamp,
        "ACCESS-PASSPHRASE": PASSPHRASE,
        "Content-Type": "application/json"
    }
    r = requests.post(url, headers=headers, data=body)
    return r.json()

# ====== Excel から金額取得 ======
def read_transfer_amounts():
    wb = load_workbook(excel_path, data_only=True)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Excelにシート {sheet_name} が存在しません")
    ws = wb[sheet_name]
    main_to_sub = ws["A2"].value or 0
    sub_to_main = ws["B2"].value or 0
    return float(main_to_sub), float(sub_to_main)

# ====== メイン処理 ======
def main():
    print("[INFO] Excelから振替金額を読み込みます...")
    main_to_sub, sub_to_main = read_transfer_amounts()
    print(f"[INFO] メイン→サブ先物: {main_to_sub}, サブ先物→メイン: {sub_to_main}")

    if main_to_sub > 0:
        print(f"[INFO] メイン Spot → サブ USDT先物 に {main_to_sub} USDT を送金します...")
        result = transfer(MAIN_UID, SUB_UID, "spot", "usdt_futures", "USDT", main_to_sub)
        print("[INFO] 振替結果(メイン→サブ):")
        print(result)

    if sub_to_main > 0:
        print(f"[INFO] サブ USDT先物 → メイン Spot に {sub_to_main} USDT を送金します...")
        result = transfer(SUB_UID, MAIN_UID, "usdt_futures", "spot", "USDT", sub_to_main)
        print("[INFO] 振替結果(サブ→メイン):")
        print(result)

if __name__ == "__main__":
    main()
