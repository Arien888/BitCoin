import sys
import io
import os
import yaml
import time
import requests
import hmac
import hashlib
import base64
import json
import openpyxl  # Excel読み込み用

# --- 出力をUTF-8に設定 ---
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")

# --- スクリプトのあるディレクトリ ---
BASE_DIR = os.path.dirname(os.path.abspath(__file__))

# --- config.yaml 読み込み ---
with open(os.path.join(BASE_DIR, "..", "config.yaml"), "r", encoding="utf-8") as f:
    config = yaml.safe_load(f)

API_KEY = config["bitget"]["api_key"]
API_SECRET = config["bitget"]["api_secret"]
PASSPHRASE = config["bitget"]["passphrase"]
IS_TESTNET = config["bitget"].get("is_testnet", False)

BASE_URL = "https://api.bitget.com" if not IS_TESTNET else "https://api-testnet.bitget.com"

# --- Excel 設定 ---
excel_rel_path = config["order_export"]["source_file"]
excel_path = os.path.join(BASE_DIR, "..", excel_rel_path)
sheet_name = config["excel"]["sheets"]["bitget_transfer"]

# --- Bitget API署名生成 ---
def sign_request(method, request_path, body=""):
    timestamp = str(time.time())  # 秒単位（小数あり）
    pre_sign = f"{timestamp}{method}{request_path}{body}"
    sign = hmac.new(API_SECRET.encode("utf-8"), pre_sign.encode("utf-8"), hashlib.sha256).digest()
    sign_b64 = base64.b64encode(sign).decode()
    return timestamp, sign_b64

# --- 資金振替 ---
def transfer_funds(from_type, to_type, coin, amount, client_oid=None):
    path = "/api/v2/spot/wallet/transfer"
    url = BASE_URL + path

    body = {
        "fromType": from_type,      # "spot"
        "toType": to_type,          # "usdt_futures"
        "coin": coin,               # "USDT"
        "amount": str(amount),      # 数値は文字列で送る
        "clientOid": client_oid or str(int(time.time() * 1000))
    }

    body_str = json.dumps(body, separators=(",", ":"))

    timestamp, sign = sign_request("POST", path, body_str)

    headers = {
        "ACCESS-KEY": API_KEY,
        "ACCESS-SIGN": sign,
        "ACCESS-TIMESTAMP": timestamp,
        "ACCESS-PASSPHRASE": PASSPHRASE,
        "Content-Type": "application/json"
    }

    r = requests.post(url, headers=headers, data=body_str)
    return r.json()

# --- Excel から金額取得 ---
def read_transfer_amounts():
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Excelにシート {sheet_name} が存在しません")
    ws = wb[sheet_name]

    spot_to_futures = ws["A2"].value or 0
    futures_to_spot = ws["B2"].value or 0

    return float(spot_to_futures), float(futures_to_spot)

# --- メイン処理 ---
def main():
    try:
        print("[INFO] Excelから振替金額を読み込みます...")
        spot_to_futures, futures_to_spot = read_transfer_amounts()
        print(f"[INFO] Spot→先物: {spot_to_futures}, 先物→Spot: {futures_to_spot}")

        if spot_to_futures > 0:
            print(f"[INFO] Spot → USDT先物に {spot_to_futures} USDT を送金します...")
            result = transfer_funds("spot", "usdt_futures", "USDT", spot_to_futures)
            print("[INFO] 振替結果(Spot→先物):")
            print(result)

        if futures_to_spot > 0:
            print(f"[INFO] USDT先物 → Spot に {futures_to_spot} USDT を送金します...")
            result = transfer_funds("usdt_futures", "spot", "USDT", futures_to_spot)
            print("[INFO] 振替結果(先物→Spot):")
            print(result)

    except Exception as e:
        import traceback
        print(f"[ERROR] 処理中に例外が発生しました: {e}")
        print(traceback.format_exc())

if __name__ == "__main__":
    main()
