import openpyxl

def write_to_existing_excel(filepath, data, keys, sheet_name):
    """
    data: dictのリスト（例: [{'asset': 'BTC', 'free': '0.01', 'locked': '0'}])
    keys: 表示するキーのリスト（例: ['asset', 'free', 'locked']）
    sheet_name: 書き込み対象のシート名
    """
    try:
        wb = openpyxl.load_workbook(filepath)
    except FileNotFoundError:
        wb = openpyxl.Workbook()

    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        # 既存データを消したければここでws.delete_rows(2, ws.max_row)
    else:
        ws = wb.create_sheet(sheet_name)

    # ヘッダー書き込み
    ws.delete_rows(1, ws.max_row)  # いったん全削除して上書きする例
    ws.append(keys)

    # データ書き込み
    for entry in data:
        row = [entry.get(k, "") for k in keys]
        ws.append(row)

    wb.save(filepath)
