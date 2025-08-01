import json
import time
import hmac
import hashlib
import requests
from openpyxl import load_workbook
from load_config import load_config  # APIキー等をyamlなどから読み込む想定

def cancel_all_orders(api_key, api_secret, symbol):
    url = "https://api.bybit.com/v5/order/cancel-all"
    timestamp = str(int(time.time() * 1000))
    recv_window = "5000"

    body = {"category": "linear", "symbol": symbol}  # USDT先物（linear）を想定
    body_json = json.dumps(body, separators=(",", ":"))

    sign_payload = f"{timestamp}{api_key}{recv_window}{body_json}"
    signature = hmac.new(api_secret.encode(), sign_payload.encode(), hashlib.sha256).hexdigest()

    headers = {
        "Content-Type": "application/json",
        "X-BAPI-API-KEY": api_key,
        "X-BAPI-TIMESTAMP": timestamp,
        "X-BAPI-RECV-WINDOW": recv_window,
        "X-BAPI-SIGN": signature,
    }

    response = requests.post(url, headers=headers, data=body_json)
    print(f"=== {symbol} キャンセルレスポンス ===")
    print(response.text)
    return response.json()

def read_cancel_list_from_excel(file_path, sheet_name, column_letter='A'):
    wb = load_workbook(file_path, data_only=True)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"シート名 '{sheet_name}' がExcelファイルに存在しません。")
    ws = wb[sheet_name]
    cancel_list = []
    for cell in ws[column_letter]:
        val = cell.value
        if val and isinstance(val, str):
            cancel_list.append(val.strip())
    return cancel_list

def main():
    config = load_config()
    api_key = config["bybit"]["key"]
    api_secret = config["bybit"]["secret"]

    # Excelファイル設定（スクリプトと同じフォルダに cancel_list.xlsx がある想定）
    excel_path = "./cancel_list.xlsx"
    sheet_name = "Orders"    # 実際のシート名に合わせて変更してください
    column_letter = "A"      # 銘柄が入っている列

    cancel_list = read_cancel_list_from_excel(excel_path, sheet_name, column_letter)

    for symbol in cancel_list:
        cancel_all_orders(api_key, api_secret, symbol)

if __name__ == "__main__":
    main()
